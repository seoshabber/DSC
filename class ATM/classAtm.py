#classAtm.py
#상속을 위한 임포트
import classBankAccount
#엑셀 파일을 열기 위한 임포트
import openpyxl
#거래 시간을 확인하기 위한 임포트
import time

class Atm(classBankAccount.BankAccount):#Atm클래스 만들기, 상속
    # 클래스 호출과 동시에 사용자와 비밀번호, 잔액을 받고, name과 balance를 받아야하는 것을 수정하기 위한 __init__함수 선언
    def __init__(self):
        self.wb_obj = openpyxl.load_workbook('UserList.xlsx')#엑셀 파일 열기
        self.sheet = self.wb_obj.active #엑셀 파일 시트 1번
        self.user_dic={}#사용자 정보 사전 생성
        for r in range(2,8):#r는 row를 의미, 반복적으로 사용자 정보를 사전에 입력하기 위한 반복문
            #사용자 사전에 추가하기, password와 balance는 정수형으로 입력
            self.user_dic[self.sheet.cell(row = r, column=1).value] = [int(self.sheet.cell(row = r, column=2).value), int(self.sheet.cell(row = r, column=3).value)]

        #print(self.user_dic)#잘 입력됬는지 확인하기

    #사용자 입력과 확인을 위한 함수 생성
    def check_user(self):
        self.user = input('Enter the username: ')#사용자 이름 입력받기
        if self.user in self.user_dic.keys():#입력한 사용자가 사전에 등록되어있는 경우
            print(self.user, '님 환영합니다.')
            #print(self.user_dic[self.user][0])  # 비밀번호 확인
            pw = int(input(self.user+'님의 비밀번호를 입력하세요: '))#비밀번호가 옳은지 확인

            if pw == self.user_dic[self.user][0]:#비밀번호를 옳게 입력한 경우
                print('사용자 정보가 확인되었습니다.')
                self.balance = self.user_dic[self.user][1]#self.deposit()과 self.withdraw()를 위한 self.balance 선언
                #print(self.user_dic)#사용자 사전 출력
                return True#참을 리턴>>참이면 메뉴 출력
            else:#비밀번호를 옳지 않게 입력한 경우
                print('사용자 정보가 정확하지 않습니다.')
                #print(self.user_dic)  # 사용자 사전 출력
                return False#거짓을 리턴>>거짓이면 메뉴가 출력되지 않음, 종료
        else:#입력한 사용자가 등록되어있지 않는 경우/추가할지 말지
            answer = input(self.user + ' 님은 등록되지 않았습니다. 추가하시겠습니까?(yes/no)')
            if answer == 'yes':#사용자를 추가하는 경우
                pw = int(input(self.user+ ' 님의 비밀번호를 입력하세요: '))#비밀번호 생성
                balance = int(input(self.user+ ' 님의 초기 잔액을 입력하세요: '))#초기 잔액 생성
                self.user_dic[self.user] = [pw, balance]#사용자 사전에 새로운 사용자 등록
                print('등록이 완료되었습니다.')
                self.balance = self.user_dic[self.user][1]#self.deposit()과 self.withdraw()를 위한 self.balance 선언
                #print(self.user_dic)#등록 확인
                return True#참을 리턴>>참이면 메뉴 출력
            else:#사용자를 추가하지 않는 경우
                print(self.user+'님을 등록하지 않습니다. 감사합니다.')
                #print(self.user_dic)  # 사용자 사전 출력
                return False#거짓을 리턴>>거짓이면 메뉴가 출력되지 않음, 종료

    def print_menu(self):#메뉴 출력하기
        print('='*30)
        print("""원하시는 메뉴를 선택하세요.
1. Deposit
2. Withdraw
3. Check Balance
4. Quit""")


    #입금을 위한 함수 선언
    def time_deposit(self):
        now_time = time.time()#명세표 발행을 위한 현재 시간 저장
        #print(self.balance)#입금 전 금액 확인
        amount=int(input('입금하실 금액을 입력하세요:'))#입금액 입력
        self.deposit(amount)#입금
        #print(self.balance)#입금 후 금액 확인
        self.user_dic[self.user][1] = self.balance#사용자 사전 갱신
        bill = input('명세표를 출력하시겠습니까?(yes/no)')#명세표 출력 여부 확인
        if bill == 'yes':#명세표를 출력하는 경우
            print('*'*50)
            print(' '*20+'명세표')
            print('거래시간:',time.asctime(time.localtime(now_time)))#전에 저장한 시간을 보기 좋게 출력
            print('이름:', self.user)
            if amount>0:#입금이 성공적으로 이루어진 경우
                print('입금액:', amount)
            else:#입금이 실패한 경우, 0원이 입금됨
                print('입금액: 0')
            print('남은 잔액:', self.user_dic[self.user][1])
            print('거래해주셔서 감사합니다. -by Python Bank')
            print('*' * 50)

        else:#명세포를 출력하지 않는 경우
            pass#아무런 일도 일어나지 않는다

    #출금을 위한 함수 선언
    def time_withdraw(self):
        now_time = time.time()#명세표 발행을 위한 현재 시간 저장
        #print(self.balance)  # 출금 전 금액 확인
        amount = int(input('출금하실 금액을 입력하세요:'))#출금액 입력
        self.withdraw(amount)#출금
        #print(self.balance)  # 출금 후 금액 확인
        self.user_dic[self.user][1] = self.balance#사용자 사전 갱신
        bill = input('명세표를 출력하시겠습니까?(yes/no)')#명세표 출력 여부 확인
        if bill == 'yes':#명세표를 출력하는 경우
            print('*' * 50)
            print(' ' * 20 + '명세표')
            print('거래시간:', time.asctime(time.localtime(now_time)))#전에 저장한 시간을 보기 좋게 출력
            print('이름:', self.user)
            if amount > self.balance:#출금이 이루어지지 않은 경우, 출금액 = 0
                print('출금액: 0')
            else:#출금이 이루어진 경우
                print('출금액:', amount)
            print('남은 잔액:', self.user_dic[self.user][1])
            print('거래해주셔서 감사합니다. -by Python Bank')
            print('*' * 50)

        else:  # 명세포를 출력하지 않는 경우
            pass  # 아무런 일도 일어나지 않는다


user = Atm()#Atm클래스 인스턴스 생성
check = user.check_user()#사용자 입력과 확인을 위해 인스턴스 함수 호출

while check == True:#사용자 확인이 끝난 경우 거래를 위한 반복문, break(종료)를 하기 전까지 반복
    #메뉴 출력
    user.print_menu()
    #메뉴 선택
    menu = int(input('>>'))
    if menu == 1:#1번 메뉴를 선택한 경우
        user.time_deposit()#입금
    elif menu == 2:#2번 메뉴를 선택한 경우
        user.time_withdraw()#출금
    elif menu == 3:#3번 메뉴를 선택한 경우
        print('현재 잔액은',user.user_dic[user.user][1],'입니다.')#잔액 확인
    elif menu == 4:#4번 메뉴를 선택한 경우
        break#종료
